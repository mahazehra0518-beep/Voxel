"""
Advanced unit and integration tests for VOXEL.
Covers Excel export, FAO/WHO motif scanning, Promiscuity indexing, and edge cases.
"""

import io
import openpyxl
import pandas as pd
import pytest
from core.sequence_parser import SequenceParser
from core.epitope_engine import EpitopeDatasetEngine, EpitopeFilteringConfig
from core.exporters import DataExporter
from models.allergenicity.allertop_adapter import AllerTOPAdapter
from models.toxicity.toxinpred_adapter import ToxinPredAdapter


def test_faowho_allergen_motif_detection():
    adapter = AllerTOPAdapter()
    # Insert known allergen motif LEEELR (Tropomyosin)
    seq_with_motif = "MKVR" + "LEEELR" + "GGAA"
    score, matches = adapter.calculate_allergenicity_index(seq_with_motif)
    assert len(matches) >= 1
    assert matches[0][0] == "LEEELR"
    assert "Tropomyosin" in matches[0][1]

    pred = adapter.predict(seq_with_motif)
    assert pred.classification == "Probable Allergen"
    assert pred.is_favorable_for_vaccine() is False


def test_toxin_motif_detection():
    adapter = ToxinPredAdapter()
    # Lytic polybasic array: KKKKKR
    seq_lytic = "MAGA" + "KKKKKR" + "AAY"
    motifs = adapter.scan_toxic_motifs(seq_lytic)
    assert len(motifs) >= 1
    assert "Polybasic" in motifs[0]

    risk = adapter.calculate_toxicity_risk_index(seq_lytic)
    assert risk > 0.40


def test_excel_export_and_multi_sheet_integrity():
    df = pd.DataFrame({
        "Epitope": ["YLQPRTFLL", "CYGVSPTKL", "WNSNNLDSK"],
        "Allele": ["HLA-A*02:01", "HLA-A*02:01", "HLA-A*01:01"],
        "IC50": [12.0, 8.0, 1200.0],
        "Rank": [0.1, 0.05, 4.0],
        "Toxicity": ["Non-Toxic", "Toxic", "Non-Toxic"],
        "Allergenicity": ["Non-Allergen", "Non-Allergen", "Non-Allergen"]
    })
    mapping = EpitopeDatasetEngine.auto_detect_columns(df)
    config = EpitopeFilteringConfig(
        enable_ic50_filter=True,
        max_ic50_nm=500.0,
        enable_rank_filter=True,
        max_binding_rank=2.0,
        exclude_toxic=True,
        exclude_allergenic=True
    )
    result = EpitopeDatasetEngine.evaluate_dataset(df, mapping, config)
    excel_bytes = DataExporter.to_excel_report(result, project_name="Test Project")

    # Read back with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    sheet_names = wb.sheetnames
    assert "Prioritized Candidates" in sheet_names
    assert "Full Audited Dataset" in sheet_names
    assert "Screening Summary" in sheet_names

    # Check prioritized tab has 1 row (+ header)
    ws_prio = wb["Prioritized Candidates"]
    assert ws_prio.max_row == 2  # 1 header + 1 passing row (YLQPRTFLL)


def test_promiscuity_and_mcda_ranking():
    # Candidate A (YLQPRTFLL) binds 3 alleles, Candidate B (KIADYNYKL) binds 1 allele
    df = pd.DataFrame({
        "Peptide": ["YLQPRTFLL", "YLQPRTFLL", "YLQPRTFLL", "KIADYNYKL"],
        "MHC": ["HLA-A*01:01", "HLA-A*02:01", "HLA-B*07:02", "HLA-A*02:01"],
        "IC50": [25.0, 30.0, 45.0, 20.0],
        "Rank": [0.2, 0.25, 0.3, 0.15],
        "Toxicity": ["Non-Toxic", "Non-Toxic", "Non-Toxic", "Non-Toxic"],
        "Allergenicity": ["Non-Allergen", "Non-Allergen", "Non-Allergen", "Non-Allergen"]
    })
    mapping = EpitopeDatasetEngine.auto_detect_columns(df)
    config = EpitopeFilteringConfig(
        weight_ic50=0.2,
        weight_rank=0.2,
        weight_antigenicity=0.0,
        weight_promiscuity=0.6  # High promiscuity weight
    )
    result = EpitopeDatasetEngine.evaluate_dataset(df, mapping, config)
    assert result.passed_candidates_count == 4
    
    # Check that rows with YLQPRTFLL have promiscuity count = 3
    prio_df = result.prioritized_df
    pep_a_rows = prio_df[prio_df["Peptide"] == "YLQPRTFLL"]
    assert (pep_a_rows["Allele_Promiscuity_Count"] == 3).all()

    pep_b_rows = prio_df[prio_df["Peptide"] == "KIADYNYKL"]
    assert (pep_b_rows["Allele_Promiscuity_Count"] == 1).all()
